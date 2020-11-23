# -*- coding: utf-8 -*-
# @Author   :   YaMeng
# @File :   lesson_7.py
# @Software :   PyCharm
# @Time :   2020/11/20 14:13
# @company  :   湖南省零檬信息技术有限公司

# 自动化的步骤：
'''
1、excel准备好测试用例，并且通过代码读取到excel的测试用例   -- read_data()
2、发送接口请求，得到响应结果    --
3、执行结果  vs  预期结果
4、写入断言结果到excel
'''
import openpyxl
import requests
import jsonpath

# 读取测试用例
def read_data(filename, sheetname):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]   #获取sheet
    max_row = sheet.max_row # 获取sheet里最大的行数
    # columns = sh.max_column  # 获取总列数
    list_1 = []  # 定义一个空列表，来接收所有的测试数据
    for i in range(2, max_row+1):  # 取头不取尾  左闭右开
        dict_1 = dict(
        id = sheet.cell(row=i, column=1).value,  # 取出id
        url = sheet.cell(row=i, column=5).value,  # 取出url
        data = sheet.cell(row=i, column=6).value,  # 取出data
        expect = sheet.cell(row=i, column=7).value)  # 取出expect
        # print(id,url,data,expect)
        list_1.append(dict_1)   # 把所有的测试数据，一一的追加到列表里
    return list_1
# 登录、注册发送请求
def api_func(url, data):
    header = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}
    res1 = requests.post(url=url, json=data, headers=header)
    response = res1.json()
    return response
# 充值、新增、提现发送请求
def api_func_1(result,url, data):
    #req_memberid = jsonpath.jsonpath(result, '$.data.id')[0]  # 获取登录用户的id   #测试用例里的充值用户id要跟登录id保持一致，不然获取不到
    req_token = jsonpath.jsonpath(result, '$.data.token_info.token')[0]  # 获取登录用户的token
    header = {"X-Lemonban-Media-Type": "lemonban.v2",
                        "Content": "application/json",
                        "Authorization": "Bearer" + " " + req_token}
    res1 = requests.post(url=url, json=data, headers=header)
    response = res1.json()
    print(response)
    return response

# 审核发送请求
def api_func_2(result,url, data):
    #req_memberid = jsonpath.jsonpath(result, '$.data.id')[0]  # 获取登录用户的id   #测试用例里的充值用户id要跟登录id保持一致，不然获取不到
    req_token = jsonpath.jsonpath(result, '$.data.token_info.token')[0]  # 获取登录用户的token
    header = {"X-Lemonban-Media-Type": "lemonban.v2",
                        "Content": "application/json",
                        "Authorization": "Bearer" + " " + req_token}
    res1 = requests.patch(url = url , json = data ,headers = header)
    # res1 = requests.post(url=url, json=data, headers=header)
    response = res1.json()
    print(response)
    return response

# 写入断言结果
def write_result(filename, sheetname, row, column, final_result):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    sheet.cell(row=row, column=column).value = final_result
    wb.save(filename)
# 执行接口自动化测试
def execute_func(filename, sheetname):
    cases = read_data(filename, sheetname)
    for case in cases:
        id = case.get('id')   # 取出id
        url = case.get('url') # 取出url
        data = case.get('data') # 取出请求参数
        expect = case.get('expect') # 取出预期结果
        expect = eval(expect)
        expect_msg = expect.get('msg')  # 取出预期结果里面的msg信息
    # 从excel取出来的数据，都是str
        data = eval(data)   #  eval()运行被字符串包裹的python表达式 ==> 字符串转换为字典
        real_result = api_func(url=url, data=data)
        real_msg = real_result.get('msg')
        # print('执行结果为：{}'.format(real_msg))
        # print('预期结果为：{}'.format(expect_msg))
        if expect_msg == real_msg:
            # print('这条测试用例通过！！')
            final_res = 'pass'
        else:
            # print('这条测试用例不通过！！！')
            final_res = 'fail'
        # print('*' * 30)
        write_result(filename, sheetname, id+1, 8, final_res)
    return real_result

def execute_func_1(filename, sheetname,result):
    cases = read_data(filename, sheetname)
    for case in cases:
        id = case.get('id')   # 取出id
        url = case.get('url') # 取出url
        data = case.get('data') # 取出请求参数
        expect = case.get('expect') # 取出预期结果
        expect = eval(expect)
        expect_msg = expect.get('msg')  # 取出预期结果里面的msg信息
    # 从excel取出来的数据，都是str
        data = eval(data)   #  eval()运行被字符串包裹的python表达式 ==> 字符串转换为字典
        real_result = api_func_1(result,url=url, data=data)
        real_msg = real_result.get('msg')
        print('执行结果为：{}'.format(real_msg))
        print('预期结果为：{}'.format(expect_msg))
        if expect_msg == real_msg:
            print('这条测试用例通过！！')
            final_res = 'pass'
        else:
            print('这条测试用例不通过！！！')
            final_res = 'fail'
        print('*' * 30)
        write_result(filename, sheetname, id+1, 8, final_res)
    return real_result

def execute_func_2(filename, sheetname,result):
    cases = read_data(filename, sheetname)
    for case in cases:
        id = case.get('id')   # 取出id
        url = case.get('url') # 取出url
        data = case.get('data') # 取出请求参数
        expect = case.get('expect') # 取出预期结果
        expect = eval(expect)
        expect_msg = expect.get('msg')  # 取出预期结果里面的msg信息
    # 从excel取出来的数据，都是str
        data = eval(data)   #  eval()运行被字符串包裹的python表达式 ==> 字符串转换为字典
        real_result = api_func_2(result,url=url, data=data)
        real_msg = real_result.get('msg')
        print('执行结果为：{}'.format(real_msg))
        print('预期结果为：{}'.format(expect_msg))
        if expect_msg == real_msg:
            print('这条测试用例通过！！')
            final_res = 'pass'
        else:
            print('这条测试用例不通过！！！')
            final_res = 'fail'
        print('*' * 30)
        write_result(filename, sheetname, id+1, 8, final_res)
    return real_result
# 管理员账户
def admin_login():
    admin_body_login = {"mobile_phone": '13124567890', "pwd": 'lemon666'}
    admin_url_login = 'http://120.78.128.25:8766/futureloan/member/login'
    admin_login_result = api_func(admin_url_login, admin_body_login)
    return admin_login_result
# 投资用户
def user_login():
    user_login_body = {"mobile_phone": '13019051897', "pwd": 'lemon666'}
    user_login_url = 'http://120.78.128.25:8766/futureloan/member/login'
    user_login_result = api_func(user_login_url, user_login_body)
    return user_login_result

# 主函数
# execute_func('test_case_api.xlsx', 'register')
real_result = execute_func('test_case_api.xlsx', 'login')
# execute_func_1('test_case_api.xlsx','recharge',real_result)
loan_add_result = execute_func_1('test_case_api.xlsx','loan_add',real_result)
# 登录管理员账户
admin_login_result = admin_login()
print(admin_login_result)
# admin_login_result是要获取管理员用户的token
execute_func_2('test_case_api.xlsx','loan_audit',admin_login_result)
# 在切换用户投资
user_login_result = user_login()
print(user_login_result)
execute_func_1('test_case_api.xlsx','withdraw',user_login_result)







